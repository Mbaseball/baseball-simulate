
from google.colab import files

# ファイルをアップロード
uploaded = files.upload()

# アップロードされたファイル名を確認
print(uploaded.keys())
batter_file_path = "Y3.xlsx"
pitcher_file_path = "P5.xlsx"
schedule_file_path = "日程.xlsx"
import pandas as pd
from google.colab import files

# ファイルをアップロード
uploaded = files.upload()

# アップロードされたファイルを使用する
batter_file_path = "Y3.xlsx"
pitcher_file_path = "P5.xlsx"
schedule_file_path = "日程.xlsx"

# Y2.xlsx（野手データ）の確認
try:
    batter_df = pd.ExcelFile(batter_file_path).parse('Sheet1')
    print("Y3.xlsx の内容:")
    print(batter_df.head())
except Exception as e:
    print(f"Y3.xlsx の読み込み中にエラーが発生: {e}")

# P4.xlsx（投手データ）の確認
try:
    pitcher_excel = pd.ExcelFile(pitcher_file_path)
    print("\nP5.xlsx のシート名:")
    print(pitcher_excel.sheet_names)

    # 投手データのサンプル表示（例: ソフトバンクシートを確認）
    pitcher_df = pitcher_excel.parse("ソフトバンク")
    print("\nP5.xlsx ソフトバンクシートの内容:")
    print(pitcher_df.head())
except Exception as e:
    print(f"P4.xlsx の読み込み中にエラーが発生: {e}")

# 日程.xlsx の確認
try:
    schedule_excel = pd.ExcelFile(schedule_file_path)
    print("\n日程.xlsx のシート名:")
    print(schedule_excel.sheet_names)

    # 日程データのサンプル表示（例: 最初のシートを確認）
    schedule_df = schedule_excel.parse(schedule_excel.sheet_names[0])
    print("\n日程.xlsx 最初のシートの内容:")
    print(schedule_df.head())
except Exception as e:
    print(f"日程.xlsx の読み込み中にエラーが発生: {e}")
# チームリスト作成
teams = []

# チーム名リスト
team_names = [
    "ソフトバンク", "楽天", "西武", "ロッテ", "日本ハム", "オリックス",
    "巨人", "阪神", "ヤクルト", "広島", "DeNA", "中日"
]

# チームデータ作成ループ
for team_name in team_names:
    try:
        # 野手データを読み込み
        batters = load_batters_from_excel(batter_file_path, team_name)

        # 投手データを読み込み
        starters, relievers, closer = load_pitchers_from_excel(pitcher_file_path, team_name)

        # チームオブジェクトを作成
        team = Team(team_name, batters, starters, relievers, closer)
        teams.append(team)
    except Exception as e:
        print(f"チーム '{team_name}' のデータ読み込みでエラー: {e}")

# 作成されたチームの確認
print(f"チームリスト: {[team.name for team in teams]}")
import pandas as pd
import random
import openpyxl

# Playerクラス
class Player:
    def __init__(self, name, stats, role, average_innings=None):
        self.name = name
        self.stats = stats
        self.role = role
        self.average_innings = average_innings  # 平均投球回数（投手のみ）
        self.reset_stats()

    def reset_stats(self):
        if self.role == "batter":
            self.hits = 0
            self.at_bats = 0
            self.walks = 0
            self.singles = 0
            self.doubles = 0
            self.triples = 0
            self.home_runs = 0
            self.rbis = 0  # 打点
        elif self.role == "pitcher":
            self.innings_pitched = 0.0
            self.runs_allowed = 0
            self.wins = 0
            self.losses = 0
            self.saves = 0
            self.appearances = 0  # 登板数

    @property
    def batting_average(self):
        return self.hits / self.at_bats if self.at_bats > 0 else 0

    @property
    def slugging_percentage(self):
        total_bases = self.singles + 2 * self.doubles + 3 * self.triples + 4 * self.home_runs
        return total_bases / self.at_bats if self.at_bats > 0 else 0

    @property
    def on_base_percentage(self):
        total_plate_appearances = self.at_bats + self.walks
        return (self.hits + self.walks) / total_plate_appearances if total_plate_appearances > 0 else 0

    @property
    def era(self):
        return (self.runs_allowed * 9) / self.innings_pitched if self.innings_pitched > 0 else 0


# Teamクラス
class Team:
    def __init__(self, name, batters, starters, relievers, closer):
        self.name = name
        self.batters = batters
        self.starters = starters
        self.relievers = relievers
        self.closer = closer
        self.starter_index = 0
        self.batter_index = 0  # 打順の現在位置を保持

    def get_starter(self):
        if not self.starters:
            raise ValueError(f"チーム {self.name} に先発投手が登録されていません。")
        starter = self.starters[self.starter_index]
        self.starter_index = (self.starter_index + 1) % len(self.starters)
        return starter

    def get_reliever(self):
        if not self.relievers:
            raise ValueError(f"チーム {self.name} に中継ぎ投手が登録されていません。")
        return random.choice(self.relievers)


# データ読み込み関数
def load_batters_from_excel(file_path, team_name):
    df = pd.ExcelFile(file_path).parse('Sheet1')
    team_data = df[df['チーム名'] == team_name].sort_values(by='打順')
    batters = []
    for _, row in team_data.iterrows():
        stats = {
            "単打": row["単打率"],
            "二塁打": row["二塁打率"],
            "三塁打": row["三塁打率"],
            "本塁打": row["本塁打率"],
            "四死球": row["出塁率"] - (row["単打率"] + row["二塁打率"] + row["三塁打率"] + row["本塁打率"])
        }
        batters.append(Player(row["選手名"], stats, "batter"))
    return batters


def load_pitchers_from_excel(file_path, sheet_name):
    df = pd.ExcelFile(file_path).parse(sheet_name, skiprows=1)
    starters = []
    relievers = []
    closer = None
    for _, row in df.iterrows():
        stats = {
            "被単打率": row["被単打率"],
            "被二塁打率": row["被二塁打率"],
            "被三塁打率": row["被三塁打率"],
            "被本塁打率": row["被本塁打率"],
            "被四死球率": row["四死球率"]
        }
        role = str(row["役割"]).strip()
        average_innings = row.get("平均投球回", None)
        player = Player(row["名前"], stats, "pitcher", average_innings)
        if "先発" in role:
            starters.append(player)
        elif "中継ぎ" in role:
            relievers.append(player)
        elif "抑え" in role:
            closer = player
    return starters, relievers, closer


def load_schedule(file_path, sheet_names):
    """
    日程エクセルファイルを読み込み、日程情報をリストで返す
    """
    schedule = []
    for sheet in sheet_names:
        df = pd.ExcelFile(file_path).parse(sheet)
        # 不要な空白を取り除く
        df["チーム1"] = df["チーム1"].str.strip()
        df["チーム2"] = df["チーム2"].str.strip()
        # 日程データをタプルで追加
        schedule.extend([(row["チーム1"], row["チーム2"], row["日程"]) for _, row in df.iterrows()])
    return schedule


# 試合シミュレーション
def simulate_game(team1, team2):
    score1, score2 = 0, 0

    # 先発投手を取得
    starter1 = team1.get_starter()
    starter2 = team2.get_starter()

    starter1_innings = random.randint(int(starter1.average_innings - 1), int(starter1.average_innings + 1))
    starter2_innings = random.randint(int(starter2.average_innings - 1), int(starter2.average_innings + 1))

    # 先発投手が投げるイニング
    for inning in range(1, 10):
        if inning <= starter1_innings:
            runs2 = simulate_inning(team2, starter1)
            starter1.runs_allowed += runs2
            starter1.innings_pitched += 1
            score2 += runs2
        else:
            reliever1 = team1.get_reliever()
            reliever1.appearances += 1
            reliever1_innings = random.randint(1, 2)
            for _ in range(reliever1_innings):
                runs2 = simulate_inning(team2, reliever1)
                reliever1.runs_allowed += runs2
                reliever1.innings_pitched += 1
                score2 += runs2

        if inning <= starter2_innings:
            runs1 = simulate_inning(team1, starter2)
            starter2.runs_allowed += runs1
            starter2.innings_pitched += 1
            score1 += runs1
        else:
            reliever2 = team2.get_reliever()
            reliever2.appearances += 1
            reliever2_innings = random.randint(1, 2)
            for _ in range(reliever2_innings):
                runs1 = simulate_inning(team1, reliever2)
                reliever2.runs_allowed += runs1
                reliever2.innings_pitched += 1
                score1 += runs1

        # 抑え投手の登板
        if inning == 9:
            if abs(score1 - score2) <= 3:  # 3点差以内なら抑えが登板
                if score1 > score2 and team1.closer:
                    closer = team1.closer
                    closer.appearances += 1
                    closer.innings_pitched += 1
                    runs2 = simulate_inning(team2, closer)
                    closer.runs_allowed += runs2
                    score2 += runs2
                elif score2 > score1 and team2.closer:
                    closer = team2.closer
                    closer.appearances += 1
                    closer.innings_pitched += 1
                    runs1 = simulate_inning(team1, closer)
                    closer.runs_allowed += runs1
                    score1 += runs1

    # 勝利投手・敗戦投手を記録
    if score1 > score2:
        starter1.wins += 1
        starter2.losses += 1
    elif score2 > score1:
        starter2.wins += 1
        starter1.losses += 1
    return score1, score2
def simulate_inning(batting_team, pitching_starter):
    runs = 0
    outs = 0
    bases = [0, 0, 0]

    while outs < 3:
        batter = batting_team.batters[batting_team.batter_index]
        batting_team.batter_index = (batting_team.batter_index + 1) % len(batting_team.batters)

        # 打席結果を決定
        result = random.choices(
            ["単打", "二塁打", "三塁打", "本塁打", "四死球", "アウト"],
            weights=[
                combine_probabilities(batter.stats.get("単打", 0), pitching_starter.stats.get("被単打率", 0)),
                combine_probabilities(batter.stats.get("二塁打", 0), pitching_starter.stats.get("被二塁打率", 0)),
                combine_probabilities(batter.stats.get("三塁打", 0), pitching_starter.stats.get("被三塁打率", 0)),
                combine_probabilities(batter.stats.get("本塁打", 0), pitching_starter.stats.get("被本塁打率", 0)),
                combine_probabilities(batter.stats.get("四死球", 0), pitching_starter.stats.get("被四死球率", 0)),
                1 - sum([
                    combine_probabilities(batter.stats.get("単打", 0), pitching_starter.stats.get("被単打率", 0)),
                    combine_probabilities(batter.stats.get("二塁打", 0), pitching_starter.stats.get("被二塁打率", 0)),
                    combine_probabilities(batter.stats.get("三塁打", 0), pitching_starter.stats.get("被三塁打率", 0)),
                    combine_probabilities(batter.stats.get("本塁打", 0), pitching_starter.stats.get("被本塁打率", 0)),
                    combine_probabilities(batter.stats.get("四死球", 0), pitching_starter.stats.get("被四死球率", 0))
                ])
            ]
        )[0]

        # 結果の処理
        if result == "アウト":
            outs += 1
            batter.at_bats += 1
        elif result == "四死球":
            batter.walks += 1
            bases, runs = advance_runners(bases, runs)
        elif result in ["単打", "二塁打", "三塁打", "本塁打"]:
            batter.at_bats += 1
            batter.hits += 1
            if result == "単打":
                batter.singles += 1
            elif result == "二塁打":
                batter.doubles += 1
            elif result == "三塁打":
                batter.triples += 1
            elif result == "本塁打":
                batter.home_runs += 1

            # 塁上ランナーの打点計算
            runs_batted_in, bases = process_hit(result, bases)
            batter.rbis += runs_batted_in
            runs += runs_batted_in

    return runs


def combine_probabilities(batter_prob, pitcher_prob):
    return (batter_prob + pitcher_prob) / 2


def advance_runners(bases, runs_scored):
    if bases[2]:
        runs_scored += 1
    bases[2] = bases[1]
    bases[1] = bases[0]
    bases[0] = 1
    return bases, runs_scored


def process_hit(hit_type, bases):
    runs = 0
    if hit_type == "単打":
        bases, runs = advance_runners(bases, runs)
    elif hit_type == "二塁打":
        runs += bases[2] + bases[1]
        bases = [0, 1, 1]
    elif hit_type == "三塁打":
        runs += sum(bases)
        bases = [0, 0, 1]
    elif hit_type == "本塁打":
        runs += sum(bases) + 1
        bases = [0, 0, 0]
    return runs, bases


def save_results_to_excel(teams, standings, results, filename="results.xlsx"):
    workbook = openpyxl.Workbook()
    standings_sheet = workbook.active
    standings_sheet.title = "リーグ成績"
    standings_sheet.append(["チーム名", "勝利数", "敗北数"])
    for team, record in standings.items():
        standings_sheet.append([team, record["wins"], record["losses"]])
    results_sheet = workbook.create_sheet("試合結果")
    results_sheet.append(["日程", "チーム1", "チーム2", "スコア1", "スコア2"])
    for result in results:
        results_sheet.append(result)
    batters_sheet = workbook.create_sheet("野手成績")
    batters_sheet.append(["チーム", "選手名", "安打数", "打率", "本塁打", "打点", "出塁率", "長打率"])
    for team in teams:
        for batter in team.batters:
            batters_sheet.append([
                team.name, batter.name, batter.hits,
                round(batter.batting_average, 3),
                batter.home_runs, batter.rbis,
                round(batter.on_base_percentage, 3),
                round(batter.slugging_percentage, 3)
            ])
    pitchers_sheet = workbook.create_sheet("投手成績")
    pitchers_sheet.append(["チーム", "選手名", "登板数", "投球回", "勝利数", "敗北数", "セーブ", "防御率"])
    for team in teams:
        for pitcher in team.starters + team.relievers + ([team.closer] if team.closer else []):
            pitchers_sheet.append([
                team.name, pitcher.name,
                pitcher.appearances, round(pitcher.innings_pitched, 1),
                pitcher.wins, pitcher.losses, pitcher.saves,
                round(pitcher.era, 2)
            ])
    workbook.save(filename)
    print(f"結果を {filename} に保存しました。")


# チーム名マッピング辞書
team_name_map = {
    "横浜": "DeNA",
    "日ハム": "日本ハム",
    "ソフトバンク": "ソフトバンク",
    "楽天": "楽天",
    "西武": "西武",
    "ロッテ": "ロッテ",
    "日本ハム": "日本ハム",
    "オリックス": "オリックス",
    "巨人": "巨人",
    "阪神": "阪神",
    "ヤクルト": "ヤクルト",
    "広島": "広島",
    "DeNA": "DeNA",
    "中日": "中日"
}


# メイン処理
def main():
    batter_file_path = "Y3.xlsx"
    pitcher_file_path = "P5.xlsx"
    schedule_file_path = "日程.xlsx"
    team_names = [
        "ソフトバンク", "楽天", "西武", "ロッテ", "日本ハム", "オリックス",
        "巨人", "阪神", "ヤクルト", "広島", "DeNA", "中日"
    ]
    teams = []
    for team_name in team_names:
        batters = load_batters_from_excel(batter_file_path, team_name)
        starters, relievers, closer = load_pitchers_from_excel(pitcher_file_path, team_name)
        team = Team(team_name, batters, starters, relievers, closer)
        teams.append(team)

    # 試合日程をロード
    sheet_names = ["3、4月", "5月", "6月", "7月", "8月", "9、10月"]
    schedule = load_schedule(schedule_file_path, sheet_names)

    # 成績の初期化
    standings = {team.name: {"wins": 0, "losses": 0} for team in teams}
    results = []

    # 試合をシミュレーション
    for game in schedule:
        team1_name = team_name_map.get(game[0].strip(), game[0].strip())
        team2_name = team_name_map.get(game[1].strip(), game[1].strip())

        team1 = next((team for team in teams if team.name == team1_name), None)
        team2 = next((team for team in teams if team.name == team2_name), None)

        if not team1 or not team2:
            print(f"エラー: チームが見つかりません - {game[0]}, {game[1]}")
            continue

        score1, score2 = simulate_game(team1, team2)
        results.append((game[2], team1.name, team2.name, score1, score2))

        # 勝敗を記録
        if score1 > score2:
            standings[team1.name]["wins"] += 1
            standings[team2.name]["losses"] += 1
        else:
            standings[team2.name]["wins"] += 1
            standings[team1.name]["losses"] += 1

    # 結果を保存
    save_results_to_excel(teams, standings, results)


if __name__ == "__main__":
    main()
